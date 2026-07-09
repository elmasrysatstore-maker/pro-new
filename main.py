import os
import glob
import tkinter as tk
from tkinter import messagebox, ttk
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tkcalendar import DateEntry

# ==================== إعدادات عامة ====================
MASTER_FILE = "الملف_الرئيسي_المحاسبي.xlsx"
EXCLUDED_FILES = {"~$", MASTER_FILE}
WINDOW_WIDTH = 700
WINDOW_HEIGHT = 500

# ألوان الواجهة
COLORS = {
    "bg": "#f5f5f5",
    "primary": "#2c3e50",
    "success": "#2ecc71",
    "danger": "#e74c3c",
    "info": "#3498db",
    "dark": "#1c2833",
    "light": "#ecf0f1"
}

# ==================== وظائف البيانات ====================

def update_customer_list():
    """تحديث قائمة العملاء من الملفات المتاحة"""
    try:
        excel_files = glob.glob("*.xlsx")
        customers = [
            os.path.splitext(f)[0]
            for f in excel_files
            if not f.startswith("~$") and f != MASTER_FILE
        ]
        customers.sort()
        combo_customer['values'] = customers
        return customers
    except Exception as e:
        messagebox.showerror("خطأ", f"فشل تحديث قائمة العملاء: {e}")
        return []

def get_all_item_types():
    """الحصول على جميع أنواع الأصناف من جميع ملفات العملاء"""
    try:
        excel_files = glob.glob("*.xlsx")
        item_types = set()
        
        for file in excel_files:
            if not file.startswith("~$") and file != MASTER_FILE:
                try:
                    df = pd.read_excel(file)
                    if "نوع الصنف" in df.columns:
                        item_types.update(df["نوع الصنف"].dropna().unique())
                except:
                    pass
        
        return sorted(list(item_types))
    except Exception as e:
        return []

def load_customer_data(event=None):
    """قراءة وعرض بيانات العميل في الجدول"""
    customer = combo_customer.get().strip()
    file_name = f"{customer}.xlsx"
    
    # مسح الجدول السابق
    for item in tree.get_children():
        tree.delete(item)
    
    if not customer or customer == MASTER_FILE:
        lbl_status.config(text="الحالة: لم يتم تحديد عميل")
        return
    
    if not os.path.exists(file_name):
        lbl_status.config(text=f"الحالة: الملف غير موجود")
        return
    
    try:
        df = pd.read_excel(file_name)
        
        # إضافة البيانات للجدول
        for _, row in df.iterrows():
            values = (
                row.get("التاريخ", ""),
                f"{row.get('الإجمالي', 0):,.2f}",
                f"{row.get('السعر', 0):,.2f}",
                row.get("الكمية", 0),
                row.get("نوع الصنف", "")
            )
            tree.insert("", tk.END, values=values)
        
        # حساب الإجماليات
        total = df["الإجمالي"].sum() if "الإجمالي" in df.columns else 0
        count = len(df)
        lbl_status.config(
            text=f"الحالة: {count} عملية | الإجمالي: {total:,.2f}"
        )
    except Exception as e:
        messagebox.showerror("خطأ", f"فشل قراءة الملف: {e}")
        lbl_status.config(text="الحالة: خطأ في القراءة")

def validate_inputs():
    """التحقق من صحة المدخلات"""
    customer = combo_customer.get().strip()
    item_type = combo_type.get().strip()
    qty_str = ent_qty.get().strip()
    price_str = ent_price.get().strip()
    
    if not customer:
        messagebox.showwarning("تحذير", "الرجاء اختيار اسم عميل!")
        return None
    
    if not item_type:
        messagebox.showwarning("تحذير", "الرجاء اختيار نوع الصنف!")
        return None
    
    if not qty_str or not price_str:
        messagebox.showwarning("تحذير", "الرجاء إدخال الكمية والسعر!")
        return None
    
    try:
        qty = float(qty_str)
        price = float(price_str)
        
        if qty <= 0 or price <= 0:
            messagebox.showerror("خطأ", "الكمية والسعر يجب أن تكون أكبر من صفر!")
            return None
            
        return {
            "customer": customer,
            "item_type": item_type,
            "qty": qty,
            "price": price,
            "total": qty * price,
            "date": date_picker.get_date().strftime('%Y-%m-%d')
        }
    except ValueError:
        messagebox.showerror("خطأ", "الكمية والسعر يجب أن تكون أرقاماً صحيحة!")
        return None

def save_data():
    """حفظ البيانات الجديدة"""
    data = validate_inputs()
    if not data:
        return
    
    file_name = f"{data['customer']}.xlsx"
    new_row = {
        "التاريخ": data['date'],
        "نوع الصنف": data['item_type'],
        "الكمية": data['qty'],
        "السعر": data['price'],
        "الإجمالي": data['total']
    }
    
    try:
        if os.path.exists(file_name):
            df = pd.read_excel(file_name)
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        else:
            df = pd.DataFrame([new_row])
        
        df.to_excel(file_name, index=False)
        messagebox.showinfo("نجاح", "تمت إضافة البيانات بنجاح!")
        
        # إعادة تحديث الواجهة
        update_customer_list()
        combo_customer.set(data['customer'])
        load_customer_data()
        
        # تحديث قائمة الأصناف
        update_item_types_list()
        
        # تجديد الحقول
        clear_inputs()
        
    except PermissionError:
        messagebox.showerror("خطأ", "الملف قيد الاستخدام! أغلق الملف أولاً.")
    except Exception as e:
        messagebox.showerror("خطأ في الحفظ", f"فشل الحفظ: {e}")

def clear_inputs():
    """مسح حقول الإدخال"""
    combo_type.set("")
    ent_qty.delete(0, tk.END)
    ent_price.delete(0, tk.END)
    date_picker.set_date(datetime.now())

def update_item_types_list():
    """تحديث قائمة أنواع الأصناف"""
    item_types = get_all_item_types()
    combo_type['values'] = item_types

def format_excel_sheet(writer, sheet_name, df):
    """تنسيق الخلايا في ملف Excel"""
    try:
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # تنسيق الرأس
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, value in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # تنسيق البيانات
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # تعديل عرض الأعمدة
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    except:
        pass

def generate_master_file():
    """إنشاء الملف الرئيسي الشامل مع الملخص"""
    excel_files = glob.glob("*.xlsx")
    valid_files = [f for f in excel_files if not f.startswith("~$") and f != MASTER_FILE]
    
    if not valid_files:
        messagebox.showwarning("تنبيه", "لا توجد ملفات عملاء لإنشاء الملف الرئيسي!")
        return
    
    summary_data = []
    
    try:
        with pd.ExcelWriter(MASTER_FILE, engine='openpyxl') as writer:
            # جمع البيانات الإحصائية
            for file in sorted(valid_files):
                customer_name = os.path.splitext(file)[0]
                df = pd.read_excel(file)
                
                summary_data.append({
                    "اسم العميل": customer_name,
                    "عدد العمليات": len(df),
                    "إجمالي الكميات": df["الكمية"].sum() if "الكمية" in df.columns else 0,
                    "إجمالي المبيعات": df["الإجمالي"].sum() if "الإجمالي" in df.columns else 0
                })
            
            # كتابة صفحة الملخص
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name="الملخص_العام", index=False)
            format_excel_sheet(writer, "الملخص_العام", df_summary)
            
            # إضافة صفحات العملاء التفصيلية
            for file in sorted(valid_files):
                customer_name = os.path.splitext(file)[0]
                df = pd.read_excel(file)
                df.to_excel(writer, sheet_name=customer_name, index=False)
                format_excel_sheet(writer, customer_name, df)
        
        messagebox.showinfo(
            "نجاح",
            f"تم تحديث الملف الرئيسي بنجاح!\n\n✓ الملخص العام\n✓ {len(valid_files)} صفحة عميل"
        )
    except PermissionError:
        messagebox.showerror("خطأ", "الملف قيد الاستخدام! أغلق الملف أولاً.")
    except Exception as e:
        messagebox.showerror("خطأ", f"فشل إنشاء الملف الرئيسي: {e}")

# ==================== بناء الواجهة ====================

root = tk.Tk()
root.title("النظام المحاسبي الذكي - v2.5")
root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}")
root.configure(bg=COLORS["bg"])
root.resizable(True, True)

# === العنوان ===
lbl_title = tk.Label(
    root,
    text="إدارة حسابات ومبيعات العملاء",
    font=("Cairo", 14, "bold"),
    bg=COLORS["bg"],
    fg=COLORS["primary"]
)
lbl_title.pack(pady=8)

# === إطار اختيار العميل ===
frame_customer = tk.LabelFrame(
    root,
    text="العميل",
    font=("Cairo", 10, "bold"),
    bg=COLORS["bg"],
    fg=COLORS["primary"],
    padx=8,
    pady=6
)
frame_customer.pack(fill="x", padx=10, pady=3)

combo_customer = ttk.Combobox(
    frame_customer,
    font=("Cairo", 10),
    justify="right",
    width=25,
    state="normal"
)
combo_customer.pack(side="right", padx=3, fill="x", expand=True)
combo_customer.bind("<<ComboboxSelected>>", load_customer_data)
combo_customer.bind("<Return>", load_customer_data)

# === إطار المدخلات المدمج ===
frame_inputs = tk.LabelFrame(
    root,
    text="إضافة صنف جديد",
    font=("Cairo", 10, "bold"),
    bg=COLORS["bg"],
    fg=COLORS["primary"],
    padx=8,
    pady=6
)
frame_inputs.pack(fill="x", padx=10, pady=3)

# الصف الأول (الصنف والكمية والسعر)
frame_row1 = tk.Frame(frame_inputs, bg=COLORS["bg"])
frame_row1.pack(fill="x", pady=2)

# الصنف
tk.Label(frame_row1, text="الصنف:", font=("Cairo", 9), bg=COLORS["bg"], width=6).pack(side="right", padx=2)
combo_type = ttk.Combobox(frame_row1, font=("Cairo", 9), justify="right", width=12, state="normal")
combo_type.pack(side="right", padx=2, fill="x", expand=False)

# الكمية
tk.Label(frame_row1, text="الكمية:", font=("Cairo", 9), bg=COLORS["bg"], width=6).pack(side="right", padx=2)
ent_qty = tk.Entry(frame_row1, font=("Cairo", 9), justify="right", width=10)
ent_qty.pack(side="right", padx=2)

# السعر
tk.Label(frame_row1, text="السعر:", font=("Cairo", 9), bg=COLORS["bg"], width=6).pack(side="right", padx=2)
ent_price = tk.Entry(frame_row1, font=("Cairo", 9), justify="right", width=10)
ent_price.pack(side="right", padx=2)

# الصف الثاني (التاريخ على اليسار)
frame_row2 = tk.Frame(frame_inputs, bg=COLORS["bg"])
frame_row2.pack(fill="x", pady=2)

# تقويم التاريخ على اليسار
tk.Label(frame_row2, text="التاريخ:", font=("Cairo", 9), bg=COLORS["bg"], width=6).pack(side="left", padx=2)
date_picker = DateEntry(
    frame_row2,
    font=("Cairo", 9),
    width=12,
    background="darkblue",
    foreground="white",
    borderwidth=2,
    year=datetime.now().year,
    month=datetime.now().month,
    day=datetime.now().day
)
date_picker.pack(side="left", padx=2)

# === أزرار العمليات ===
frame_buttons = tk.Frame(root, bg=COLORS["bg"])
frame_buttons.pack(fill="x", padx=10, pady=3)

tk.Button(
    frame_buttons,
    text="💾 حفظ",
    font=("Cairo", 10, "bold"),
    bg=COLORS["success"],
    fg="white",
    command=save_data,
    width=20
).pack(side="left", padx=2, fill="x", expand=True)

tk.Button(
    frame_buttons,
    text="🗑️ مسح",
    font=("Cairo", 10, "bold"),
    bg=COLORS["light"],
    fg=COLORS["primary"],
    command=clear_inputs,
    width=20
).pack(side="left", padx=2, fill="x", expand=True)

tk.Button(
    frame_buttons,
    text="📊 تحديث الملف",
    font=("Cairo", 10, "bold"),
    bg=COLORS["dark"],
    fg="white",
    command=generate_master_file,
    width=20
).pack(side="left", padx=2, fill="x", expand=True)

# === الجدول ===
frame_table = tk.LabelFrame(
    root,
    text="بيانات العميل",
    font=("Cairo", 10, "bold"),
    bg=COLORS["bg"],
    fg=COLORS["primary"]
)
frame_table.pack(fill="both", expand=True, padx=10, pady=3)

columns = ("التاريخ", "الإجمالي", "السعر", "الكمية", "الصنف")
tree = ttk.Treeview(frame_table, columns=columns, show="headings", height=6)

# تكوين الأعمدة
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, anchor="center", width=95)

tree.pack(fill="both", expand=True, padx=3, pady=3)

# شريط التمرير
scrollbar = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
scrollbar.pack(side="right", fill="y")
tree.configure(yscrollcommand=scrollbar.set)

# === شريط الحالة ===
lbl_status = tk.Label(
    root,
    text="الحالة: جاهز",
    font=("Cairo", 9),
    bg=COLORS["light"],
    fg=COLORS["primary"],
    relief="sunken",
    anchor="w"
)
lbl_status.pack(fill="x", padx=5, pady=3)

# === تشغيل البرنامج ===
update_customer_list()
update_item_types_list()

root.mainloop()
